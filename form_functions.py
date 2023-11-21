import sqlite3
import datetime
from wtforms import SubmitField
import shutil
import numpy
import io
import os
import tempfile
import xlwings as xw
import pandas
from io import BytesIO
from flask import make_response
from openpyxl import Workbook
from openpyxl import load_workbook
import difflib
import datetime
from wtforms import StringField, PasswordField, SubmitField, BooleanField, SelectField, FieldList, FormField, DecimalField, DateField, SelectMultipleField, IntegerField, HiddenField
import openpyxl.utils.dataframe as dataframe
from flask import flash
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
import datetime


'''class DynamicSelectField(SelectField):
    def update_choices(self, choices):
        self.choices = choices'''

db_file = "data_base.db"

def generate_user_filling_timesheet_report(start_date, end_date):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    header_array = ['username']
    current_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    while current_date <= datetime.datetime.strptime(end_date, "%Y-%m-%d"):
        header_array.append(current_date.strftime("%Y-%m-%d"))
        current_date = current_date + datetime.timedelta(days=1)

    cursor.execute("SELECT user_name from user")
    usernames_query = cursor.fetchall()
    usernames = [item[0] for item in usernames_query]
    table_value = []
    color_value = []
    table_value.append(header_array)
    color_value.append([''] * len(header_array))

    for username in usernames:
        row_color_array = ['']
        row_value_array = [username]
        for n in range(1, len(header_array), 1):
            color, _, _, status = color_coding(datetime.datetime.strptime(header_array[n], '%Y-%m-%d'), username)
            print(f'status = {status}')
            row_color_array.append(color)
            row_value_array.append(status)
        color_value.append(row_color_array)
        table_value.append(row_value_array)
    wb = openpyxl.Workbook()
    ws = wb.active

    color_code_dict = {'yellow': 'FFFF00', 'green': '00FF00', 'red': 'FF0000', 'orange': 'FFA500', 'white': 'FFFFFF', '': 'FFFFFF'}

    for rowIndex in range(0, len(table_value), 1):
        for colIndex in range(0, len(table_value[rowIndex]), 1):
            color_code = color_code_dict[color_value[rowIndex][colIndex]]
            print('filling excel cells')
            fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
            cell = ws.cell(row=rowIndex + 1, column=colIndex + 1, value=table_value[rowIndex][colIndex])
            cell.fill = fill
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                            bottom=Side(style='thin'))
            cell.border = border

    cursor.close()
    connection.close()

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    response = make_response(excel_data.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=my_excel_file.xlsx'

    return response

def get_query_data(query):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()
    connection.close()
    return data

def get_access_privilege(username):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'SELECT access_privilege FROM user WHERE user_name = "{username}"')
    access_string = cursor.fetchone()[0]
    print(access_string)
    if type(access_string) == str:
        access_list = access_string.split(',')
    else:
        access_list = []
    modified_list = []
    for item in access_list:
        modified_list.append(item.strip())
    cursor.close()
    connection.close()
    return modified_list

def update_download_log(status):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    if status == 'downloaded':
        cursor.execute(f'INSERT INTO download_log (status) VALUES (?)', (status,))
    elif status == 'uploaded':
        query = "SELECT MAX(entry) FROM download_log"
        cursor.execute(query)
        max_primary_key = cursor.fetchone()[0]
        cursor.execute(f'UPDATE download_log SET status = "uploaded" WHERE entry = {max_primary_key}')
    connection.commit()
    cursor.close()
    connection.close()


def check_download_log(entry_number, status):
    error = False
    message = 'success'
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    query = "SELECT MAX(entry) FROM download_log"
    cursor.execute(query)
    max_primary_key = cursor.fetchone()[0]
    if max_primary_key is not None:
        cursor.execute(f'SELECT status FROM download_log WHERE entry = {max_primary_key}')
        last_status = cursor.fetchone()[0]
    else:
        max_primary_key = 0
        last_status = None

    if max_primary_key == 0:
        error = False
        message = 'success'
        #cursor.execute(f'INSERT INTO download_log (entry, status) VALUES (?, ?)', (int(max_primary_key) + 1, status))

    elif int(max_primary_key) > 0 and status == 'downloaded':
        #if last_status == 'downloaded':
            #message = f'Excel file "database_{max_primary_key}" has been downloaded. Please upload it before downloading to avoid data conflict'
        message = 'success'
        error = False
        #cursor.execute(f'INSERT INTO download_log (entry, status) VALUES (?, ?)', (int(max_primary_key) + 1, status))

        '''elif last_status == 'uploaded':
            error = False
            message = 'Success'
            cursor.execute(f'INSERT INTO download_log (entry, status) VALUES (?)', (max_primary_key + 1, status))'''

    elif int(max_primary_key) > 0 and status == 'uploaded':
        print('uploading')
        print(f'max key = {max_primary_key}')
        print(f'file entry = {entry_number}')

        if int(entry_number) == int(max_primary_key):
            print('max key == file entry')
            print(f'last status = {last_status}')
            if last_status == 'downloaded':
                message = 'success'
                error = False
                #print(f'UPDATE download_log SET status = {status} WHERE entry = {max_primary_key}')
                #cursor.execute(f'UPDATE download_log SET status = {status} WHERE entry = {max_primary_key}')
            elif last_status == 'uploaded':
                print('already uploaded before')
                message = 'Data from this Excel file has already be uploaded before. If more changes need to be made, please download new Excel file'
                error = True

        else:
            error = True
            print('file not last version')
            message = f'This Excel file is not the latest downloaded version. The last downloaded Excel file is "database_{max_primary_key}"'




    connection.commit()
    cursor.close()
    connection.close()

    return message, error, max_primary_key



def find_column_datatype(tablename, header):
    target_column = tablename + '>>' + header
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute('PRAGMA TABLE_INFO(datatype_setting)')
    header_info = cursor.fetchall()
    headers = [row[1] for row in header_info]
    print(f'SELECT * FROM datatype_setting WHERE column_reference = {target_column}')
    cursor.execute(f'SELECT * FROM datatype_setting WHERE column_reference = "{target_column}"')
    selected_row = cursor.fetchone()
    if selected_row is not None:
        datatype_index = headers.index('datatype')
        datatype = selected_row[datatype_index]
    else:
        datatype = 'text'
    return datatype



def update_password(user_name, new_password):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'UPDATE user SET password = "{new_password}" WHERE user_name = "{user_name}"')
    connection.commit()
    cursor.close()
    connection.close()

def find_email_by_username(usernames):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    emails = []
    for username in usernames:
        cursor.execute(f'SELECT email FROM user WHERE user_name = "{username}"')
        email = cursor.fetchone()[0]
        emails.append(email)
    cursor.close()
    connection.close()
    return emails

def find_username_by_email(email):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'SELECT user_name FROM user WHERE email = "{email}"')
    username = cursor.fetchone()[0]
    print(f'found username = {username}')
    cursor.close()
    connection.close()
    return username

def rename_header(table_name, original_header, new_header):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    header_info = cursor.fetchall()
    headers = [row[1] for row in header_info]
    if original_header in headers:
        print(f'ALTER TABLE {table_name} RENAME COLUMN {original_header} TO {new_header}')
        cursor.execute(f'ALTER TABLE {table_name} RENAME COLUMN {original_header} TO {new_header}')
        connection.commit()
    cursor.close()
    connection.close()

def remove_header(table_name, header):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    header_info = cursor.fetchall()
    headers = [row[1] for row in header_info]

    if header in headers:
        cursor.execute(f'ALTER TABLE {table_name} DROP COLUMN {header}')
        connection.commit()
    cursor.close()
    connection.close()

def get_database_headers(table_name):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    header_info = cursor.fetchall()
    cursor.close()
    connection.close()
    return [row[1] for row in header_info]

def get_header_index(table_name, headers):
    header_index_array = []
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    header_info = cursor.fetchall()
    cursor.close()
    connection.close()
    headers_from_database = [row[1] for row in header_info]
    for header in headers:
        header_index_array.append(headers_from_database.index(header))

    return header_index_array

def select_field_type(table_name, column_name):
    dropdown_table_name = 'dropdown_list_setting'
    target_name = table_name + '>>' + column_name
    dropdown_headers = ['dropdown_list_type']
    dropdown_table_content = download_table(dropdown_table_name, dropdown_headers, conditions=[['target', '=', target_name]])
    type_list = []

    for row in dropdown_table_content:
        if row[0] in ['single_select', 'multi_select']:
            type_list.append(row[0])
    print(f'target name = {target_name}')
    print(type_list)
    if len(type_list) > 0 and all('single_select' == element for element in type_list):
        return 'single_select'
    elif len(type_list) and all('multi_select' == element for element in type_list):
        return 'multi_select'
    else:
        return None

def list_in_list(sub_list, main_list):

    truths = []
    for item in sub_list:
        if item in main_list:
            truths.append(True)
        else:
            truths.append(False)
    if len(truths) > 0 and all(truths):
        return True
    else:
        return False

def is_select_field(table_name, column_name):
    dropdown_table_name = 'dropdown_list_setting'
    target_name = table_name + '>>' + column_name
    dropdown_headers = ['target', 'source']
    dropdown_table_content = download_table(dropdown_table_name, dropdown_headers)
    for n in range(1, len(dropdown_table_content), 1):
        if dropdown_table_content[n][0] == target_name and dropdown_table_content[n][1] not in ['', None]:
            return True
    return False

def get_dropdown_list(table_name, column_name, entry_index, headers=[], cell_values=[]):

    if 'entry' in headers:
        entry_index = cell_values[headers.index('entry')]

    condition = [['entry', '=', entry_index]]
    dropdown_table_name = 'dropdown_list_setting'
    print(f'table name = {table_name}, column_name = {column_name}')
    target_name = table_name + '>>' + column_name
    dropdown_headers = ['target', 'source', 'criteria_1', 'filter_1', 'criteria_2', 'filter_2', 'criteria_3', 'filter_3', 'criteria_4', 'filter_4']
    dropdown_table_content = download_table(dropdown_table_name, dropdown_headers)

    target_col = dropdown_headers.index('target')
    source_col = dropdown_headers.index('source')
    filter_1_col = dropdown_headers.index('filter_1')
    filter_2_col = dropdown_headers.index('filter_2')
    filter_3_col = dropdown_headers.index('filter_3')
    filter_4_col = dropdown_headers.index('filter_4')
    criteria_1_col = dropdown_headers.index('criteria_1')
    criteria_2_col = dropdown_headers.index('criteria_2')
    criteria_3_col = dropdown_headers.index('criteria_3')
    criteria_4_col = dropdown_headers.index('criteria_4')


    dropdown_list = []

    print(f'length of dropdown_table_content = {len(dropdown_table_content)}')
    for row in range(1, len(dropdown_table_content), 1):

        if dropdown_table_content[row][target_col] == target_name:
            source_cell = dropdown_table_content[row][source_col]
            if '>>' in source_cell:
                source_table_name = source_cell.split('>>')[0]
                source_table_column = source_cell.split('>>')[1]
                source_column = download_table(source_table_name, [source_table_column])
                source_column = [source_column[n][0] for n in range(1, len(source_column), 1)]
            elif ',' in source_cell:
                source_column = [substring.strip() for substring in source_cell.split(',')]


            filter_list = []
            criteria_list = []

            if dropdown_table_content[row][filter_1_col] is not None and dropdown_table_content[row][criteria_1_col] is not None:

                filter_1_table = dropdown_table_content[row][filter_1_col].split('>>')[0]
                filter_1_column = dropdown_table_content[row][filter_1_col].split('>>')[1]
                filter_list_temp = download_table(filter_1_table, [filter_1_column])
                filter_list_temp = [filter_list_temp[n][0] for n in range(1, len(filter_list_temp), 1)]
                filter_list.append(filter_list_temp)

                criteria_1_table = dropdown_table_content[row][criteria_1_col].split('>>')[0]
                criteria_1_column = dropdown_table_content[row][criteria_1_col].split('>>')[1]
                if type(condition[0][2]) in [int, float]:
                    criteria_list_temp = download_table(criteria_1_table, [criteria_1_column], conditions=condition)
                    criteria_list_temp = criteria_list_temp[1][0]
                    if type(criteria_list_temp) == str:
                        criteria_list_temp = criteria_list_temp.split(',')
                    else:
                        criteria_list_temp = []
                else:
                    criteria_list_temp = []

                if criteria_1_table == table_name and criteria_1_column in headers:
                    print(f'headers = {headers}')
                    print(f'cell values = {cell_values}')
                    criteria_index = headers.index(criteria_1_column)
                    print(f'criteria index = {criteria_index}')
                    criteria_list_temp = cell_values[criteria_index]
                criteria_list.append(criteria_list_temp)

            if dropdown_table_content[row][filter_2_col] is not None and dropdown_table_content[row][criteria_2_col] is not None:

                filter_2_table = dropdown_table_content[row][filter_2_col].split('>>')[0]
                filter_2_column = dropdown_table_content[row][filter_2_col].split('>>')[1]
                filter_list_temp = download_table(filter_2_table, [filter_2_column])
                filter_list_temp = [filter_list_temp[n][0] for n in range(1, len(filter_list_temp), 1)]
                filter_list.append(filter_list_temp)

                criteria_2_table = dropdown_table_content[row][criteria_2_col].split('>>')[0]
                criteria_2_column = dropdown_table_content[row][criteria_2_col].split('>>')[1]
                if type(condition[0][2]) in [int, float]:
                    criteria_list_temp = download_table(criteria_2_table, [criteria_2_column], conditions=condition)
                    criteria_list_temp = criteria_list_temp[1][0]
                    if type(criteria_list_temp) == str:
                        criteria_list_temp = criteria_list_temp.split(',')
                    else:
                        criteria_list_temp = []
                else:
                    criteria_list_temp = []
                if criteria_2_table == table_name and criteria_2_column in headers:
                    criteria_index = headers.index(criteria_2_column)
                    criteria_list_temp = cell_values[criteria_index]
                criteria_list.append(criteria_list_temp)


            if dropdown_table_content[row][filter_3_col] is not None and dropdown_table_content[row][criteria_3_col] is not None:

                filter_3_table = dropdown_table_content[row][filter_3_col].split('>>')[0]
                filter_3_column = dropdown_table_content[row][filter_3_col].split('>>')[1]
                filter_list_temp = download_table(filter_3_table, [filter_3_column])
                filter_list_temp = [filter_list_temp[n][0] for n in range(1, len(filter_list_temp), 1)]
                filter_list.append(filter_list_temp)

                criteria_3_table = dropdown_table_content[row][criteria_3_col].split('>>')[0]
                criteria_3_column = dropdown_table_content[row][criteria_3_col].split('>>')[1]
                if type(condition[0][2]) in [int, float]:
                    criteria_list_temp = download_table(criteria_3_table, [criteria_3_column], conditions=condition)
                    criteria_list_temp = criteria_list_temp[1][0]
                    if type(criteria_list_temp) == str:
                        criteria_list_temp = criteria_list_temp.split(',')
                    else:
                        criteria_list_temp = []
                else:
                    criteria_list_temp = []
                if criteria_3_table == table_name and criteria_3_column in headers:
                    criteria_index = headers.index(criteria_3_column)
                    criteria_list_temp = cell_values[criteria_index]
                criteria_list.append(criteria_list_temp)

            if dropdown_table_content[row][filter_4_col] is not None and dropdown_table_content[row][criteria_4_col] is not None:

                filter_4_table = dropdown_table_content[row][filter_4_col].split('>>')[0]
                filter_4_column = dropdown_table_content[row][filter_4_col].split('>>')[1]
                filter_list_temp = download_table(filter_4_table, [filter_4_column])
                filter_list_temp = [filter_list_temp[n][0] for n in range(1, len(filter_list_temp), 1)]
                filter_list.append(filter_list_temp)

                criteria_4_table = dropdown_table_content[row][criteria_4_col].split('>>')[0]
                criteria_4_column = dropdown_table_content[row][criteria_4_col].split('>>')[1]
                if type(condition[0][2]) in [int, float]:
                    criteria_list_temp = download_table(criteria_4_table, [criteria_4_column], conditions=condition)
                    criteria_list_temp = criteria_list_temp[1][0]
                    if type(criteria_list_temp) == str:
                        criteria_list_temp = criteria_list_temp.split(',')
                    else:
                        criteria_list_temp = []
                else:
                    criteria_list_temp = []
                if criteria_4_table == table_name and criteria_4_column in headers:
                    criteria_index = headers.index(criteria_4_column)
                    criteria_list_temp = cell_values[criteria_index]
                criteria_list.append(criteria_list_temp)

            dropdown_list_temp = []

            if len(filter_list) > 0:
                for n in range(0, len(filter_list), 1):
                    dropdown_list_temp.append([])
                    for m in range(0, len(filter_list[n]), 1):
                        #print(f'criteria_list = {criteria_list} and filter_list[n][m] = {filter_list[n][m]}')
                        if filter_list[n][m] is not None and criteria_list[n] is not None and list_in_list(criteria_list[n], filter_list[n][m].split(',')):
                            if source_column[m] not in dropdown_list:
                                dropdown_list_temp[-1].append(source_column[m])
            else:
                dropdown_list_temp.append([])
                for m in range(0, len(source_column), 1):
                    if source_column[m] not in dropdown_list:
                        dropdown_list_temp[-1].append(source_column[m])


            for n in range(0, len(dropdown_list_temp), 1):
                for m in range(0, len(dropdown_list_temp[n]), 1):
                    if dropdown_list_temp[n][m] is not None and dropdown_list_temp[n][m] not in dropdown_list:
                        truth_vector = [dropdown_list_temp[n][m] in dropdown_list_temp[k] for k in range(0, len(dropdown_list_temp), 1)]
                        if all(truth_vector):
                            dropdown_list.append(dropdown_list_temp[n][m])
    dropdown_list.insert(0, '')

    return dropdown_list

def populate_excel(workbook, worksheet, data_table):
    ws = workbook.create_sheet(worksheet)
    fill1 = PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type='solid')
    fill2 = PatternFill(start_color='80ADD8E6', end_color='80ADD8E6', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_bold_font = Font(bold=True)
    header_border = Border(bottom=Side(style='thick'))
    body_bold_font = Font(bold=False)
    body_border = Border(bottom=Side(style='dotted'))
    for row in range(0, len(data_table), 1):
        for col in range(0, len(data_table[row]), 1):
            if row == 0 and worksheet not in ['column_table', 'row_table', 'dropdown_list_setting', 'datatype_setting']:
                ws.cell(row=row + 1, column=col + 1, value=data_table[row][col].replace('_', ' '))

            else:
                ws.cell(row=row + 1, column=col + 1, value=data_table[row][col])

            ws.cell(row=row + 1, column=col + 1, value=data_table[row][col]).alignment = alignment
            column_letter = ws.cell(row=row + 1, column=col + 1).column_letter
            ws.column_dimensions[column_letter].auto_size = False
            ws.column_dimensions[column_letter].width = 20
            if col % 2 == 0:
                ws.cell(row=row + 1, column=col + 1).fill = fill1
            else:
                ws.cell(row=row + 1, column=col + 1).fill = fill2
            if row == 0:
                ws.cell(row=row + 1, column=col + 1).font = header_bold_font
                ws.cell(row=row + 1, column=col + 1).border = header_border
            else:
                ws.cell(row=row + 1, column=col + 1).font = body_bold_font
                ws.cell(row=row + 1, column=col + 1).border = body_border

    if worksheet not in ['column_table', 'row_table']:
        ws.column_dimensions['A'].hidden = True
    else:
        ws.sheet_state = "hidden"

    '''worksheet = workbook['Introduction']
    for shape in worksheet.shapes:
        worksheet.add_shape(shape)'''
    ws.auto_filter.ref = ws.dimensions


def download_table(table_name, headers, conditions = None):

    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    headers_string = ','.join(headers)

    if conditions is None:
        query = f'SELECT {headers_string} FROM {table_name}'
    else:
        conditions_string = query_conditions(conditions)
        query = f'SELECT {headers_string} FROM {table_name} WHERE {conditions_string}'
    print(query)
    cursor.execute(query)
    body = cursor.fetchall()

    return [tuple(headers)] + body

def convert_header_to_column(table_name, header):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    result = cursor.fetchall()
    headers = [row[1] for row in result]
    if header in headers:
        column = headers.index(header) + 1
    else:
        column = 0
    cursor.close()
    connection.close()

    return column_number_to_letter(column)

def column_number_to_letter(column_number):
    letter = ''
    while column_number > 0:
        column_number, remainder = divmod(column_number - 1, 26)
        letter = chr(65 + remainder) + letter
    return '$' + letter

def get_database_headers_info(table_name_array):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    header_info_array = []

    for table_name in table_name_array:
        cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
        header_info = cursor.fetchall()
        header_info_array.append(header_info)

    cursor.close()
    connection.close()

    return header_info_array

def color_coding_approval(submit_date, user_name=None):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()

    table_name = 'time_sheet'
    if user_name in [None, '']:
        conditions = [['submit_date', '=', submit_date]]
    else:
        conditions = [['user_name', '=', user_name], ['submit_date', '=', submit_date]]
    query_string = query_conditions(conditions)

    cursor.execute(f'SELECT status, hours FROM {table_name} WHERE {query_string}')
    fetched_data = cursor.fetchall()
    status_array = [fetched_data[n][0] for n in range(0, len(fetched_data), 1)]
    hours_array = [fetched_data[n][1] for n in range(0, len(fetched_data), 1)]

    status = ''
    filtered_indices = [index for index in range(0, len(status_array), 1) if status_array[index] in ['submitted', 'rejected', 'approved']]
    filtered_status_array = [status_array[n] for n in filtered_indices]

    filtered_hours_array = [hours_array[n] for n in filtered_indices]
    print(f'submit date = {submit_date} and filtered status array = {filtered_status_array}')

    for n in range(0, len(filtered_hours_array), 1):
        if filtered_hours_array[n] in ['', None]:
            filtered_hours_array[n] = 0

    color = 'white'
    if len(filtered_status_array) > 0:
        if all([element == 'approved' for element in filtered_status_array]):
            status = 'approved'
            color = 'green'
        elif 'rejected' in filtered_status_array:
            status = 'rejected'
            color = 'red'
        elif 'submitted' in filtered_status_array:
            status = 'submitted'
            color = 'yellow'


    total_hours = sum(filtered_hours_array)

    #submitted_hours = sum([hours_array[n] for n in range(0, len(status_array), 1) if status_array[n] in ['submitted', 'rejected']])
    approved_hours = sum([filtered_hours_array[n] for n in range(0, len(filtered_status_array), 1) if filtered_status_array[n] == 'approved'])

    connection.commit()
    cursor.close()
    connection.close()

    return color, total_hours, approved_hours, status

def color_coding(submit_date, user_name=None):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()

    table_name = 'time_sheet'
    if user_name in [None, '']:
        conditions = [['submit_date', '=', submit_date]]
    else:
        conditions = [['user_name', '=', user_name], ['submit_date', '=', submit_date]]
    query_string = query_conditions(conditions)

    cursor.execute(f'SELECT status, hours FROM {table_name} WHERE {query_string}')
    fetched_data = cursor.fetchall()
    status_array = [fetched_data[n][0] for n in range(0, len(fetched_data), 1)]
    print(f'status array = {status_array}')
    hours_array = [fetched_data[n][1] for n in range(0, len(fetched_data), 1)]
    print(f'hours array = {hours_array}')
    for n in range(0, len(hours_array), 1):
        if hours_array[n] is None:
            hours_array[n] = 0
    total_hours = sum(hours_array)
    ##############
    status = ''
    if len(status_array) > 0:
        if all([stat == 'approved' for stat in status_array]):
            status = 'approved'
        elif 'rejected' in status_array:
            status = 'rejected'
        elif 'saved' not in status_array:
            status = 'submitted'
        elif 'rejected' not in status_array and 'saved' in status_array:
            status = 'saved'

    ################
    #submitted_hours = sum([hours_array[n] for n in range(0, len(status_array), 1) if status_array[n] in ['submitted', 'rejected']])
    approved_hours = sum([hours_array[n] for n in range(0, len(status_array), 1) if status_array[n] == 'approved'])
    color = 'white'
    '''if len(status_array) == 0:
        color = 'white'
    elif all(status == 'approved' for status in status_array):
        color = 'green'
    elif all(status == 'submitted' for status in status_array):
        color = 'yellow'
    elif 'rejected' not in status_array and 'saved' in status_array:
        color = 'orange'
    elif 'rejected' in status_array:
        color = 'red'''
    if status == '':
        color = 'white'
    elif status == 'approved':
        color = 'green'
    elif status == 'submitted':
        color = 'yellow'
    elif status == 'saved':
        color = 'orange'
    elif status == 'rejected':
        color = 'red'

    connection.commit()
    cursor.close()
    connection.close()

    return color, total_hours, approved_hours, status

def save_show_list(table_name, show_list):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()

    if 'password' in show_list:
        show_list.remove('password')

    show_list_string = ','.join(show_list)

    query = f'update table_metadata set show_in_summary = "{show_list_string}" where table_name = "{table_name}"'

    cursor.execute(query)
    connection.commit()
    cursor.close()
    connection.close()

def get_column_list(table_name):

    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()

    cursor.execute(f'SELECT show_in_summary FROM table_metadata WHERE table_name = "{table_name}"')

    show_list = cursor.fetchone()
    if show_list != None:
        show_list = show_list[0]
        show_list_array = show_list.split(',')
    else:
        show_list_array = []

    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    header_info = cursor.fetchall()
    full_header_list = [row[1] for row in header_info]

    cursor.close()
    connection.close()

    if 'password' in show_list_array:
        show_list_array.remove('password')

    if 'password' in full_header_list:
        full_header_list.remove('password')


    return show_list_array, full_header_list

def save_uploaded(table_name, uploaded_data):
    print(f'table_name = {table_name}')

    new_data = [list(row) for row in uploaded_data]
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers_info = cursor.fetchall()
    database_headers = [row[1] for row in database_headers_info]
    save_headers = new_data[0]
    save_headers = [save_header.replace(' ', '_') for save_header in save_headers]
    save_data = new_data[1:]
    mod_data_type = []

    ####################### Deleted columns ###############################

    n = 0
    print(f'saved headers = {save_headers}')
    while n < len(save_headers):

        if '$delete' in save_headers[n] and save_headers[n].split('$')[0] in [column.split('$')[0] for column in database_headers]:
            index = [column.split('$')[0] for column in database_headers].index(save_headers[n].split('$')[0])
            cursor.execute(f'ALTER TABLE {table_name} DROP COLUMN {database_headers[index]}')
            save_headers.pop(n)
            for m in range(0, len(save_data), 1):
                save_data[m].pop(n)
            n -= n
        n += 1

    ######################### Renamed columns ###########################################

    for n in range(0, len(save_headers), 1):
        if len(save_headers[n].split('$')) > 1:
            attributes = save_headers[n].split('$')[1:]

            if [('rename' in attribute) for attribute in attributes].count(True) and save_headers[n].split('$')[0] in database_headers:
                index = [('rename' in attribute) for attribute in attributes].index(True)
                primary_name = save_headers[n].split('$')[index + 1].replace('rename', '').replace('(', '').replace(')', '')
                attributes.pop(index)
            else:
                primary_name = save_headers[n].split('$')[0]
            attributes_string = '$'.join(attributes)
            full_rename_string = primary_name + attributes_string
            index = [column.split('$')[0] for column in database_headers].index(save_headers[n].split('$')[0])
            cursor.execute(f'ALTER TABLE {table_name} RENAME COLUMN {database_headers[index]} TO {full_rename_string}')
            save_headers[n] = full_rename_string

    ########################################################################

    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers_info = cursor.fetchall()
    database_headers = [row[1] for row in database_headers_info]

    mod_headers = database_headers.copy()
    ######################## Rearrange columns###########################

    for n in range(0, len(save_headers), 1):

        if n > 0:
            if save_headers[n] not in mod_headers:
                insert_index = mod_headers.index(save_headers[n - 1]) + 1
                mod_headers.insert(insert_index, save_headers[n])
            elif mod_headers.index(save_headers[n]) < mod_headers.index(save_headers[n - 1]):
                mod_headers.pop(mod_headers.index(save_headers[n]))
                insert_index = mod_headers.index(save_headers[n - 1]) + 1
                mod_headers.insert(insert_index, save_headers[n])
        else:
            if save_headers[n] not in mod_headers:
                insert_index = 0
                mod_headers.insert(insert_index, save_headers[n])

    for header in mod_headers:
        if header in database_headers:
            index = database_headers.index(header)
            #mod_data_type.append(database_headers_info[index][2])
            if header == 'entry':
                mod_data_type.append('INTEGER PRIMARY KEY')
            else:
                mod_data_type.append(database_headers_info[index][2])
        else:
            index = save_headers.index(header)
            if header == 'entry':
                mod_data_type.append('INTEGER PRIMARY KEY')
            elif is_number(save_data[0][index]):
                mod_data_type.append('REAL')
            elif is_date(save_data[0][index]):
                mod_data_type.append('DATE')
            else:
                mod_data_type.append('TEXT')

    header_query_array = []
    for n in range(0, len(mod_headers), 1):
        header_query_array.append(mod_headers[n] + ' ' + mod_data_type[n])

    header_query_string = ', '.join(header_query_array)
    if mod_headers != database_headers:
        temp_table = table_name + '_temp'
        cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{temp_table}'")
        result = cursor.fetchone()
        if result is not None:
            cursor.execute(f'DROP TABLE {temp_table}')

        cursor.execute(f'CREATE TABLE {temp_table} ({header_query_string})')

        common_headers = []
        for header in mod_headers:
            if header in database_headers:
                common_headers.append(header)

        select_column_string = ', '.join(common_headers)
        query = f'SELECT {select_column_string} FROM {table_name}'
        print(query)
        print(table_name)
        cursor.execute(query)
        copied_table = cursor.fetchall()

        insert_column_string = ', '.join([header for header in common_headers])
        value_string = ', '.join(['?'] * len(common_headers))

        query = f'INSERT INTO {temp_table} ({insert_column_string}) VALUES ({value_string})'

        cursor.execute(f'PRAGMA table_info({temp_table})')

        cursor.executemany(query, copied_table)

        cursor.execute(f'DROP TABLE {table_name}')
        cursor.execute(f'ALTER TABLE {temp_table} RENAME TO {table_name}')

    ##############################delete empty rows ########################################
    n = 0
    if 'entry' in save_headers:
        entry_index = save_headers.index('entry')
        while n < len(save_data):
            empty = False
            m = 0
            while True:
                if m == entry_index:
                    m += 1
                elif save_data[n][m] in ['', None]:
                    m += 1
                else:
                    break

                if m >= len(save_data[n]):
                    empty = True
                    break
            if empty:
                entry = save_data[n][entry_index]
                save_data.pop(n)
                cursor.execute(f'DELETE FROM {table_name} WHERE entry = {entry}')
                n -= 1
            n += 1

    ###############################inserting new data#######################################
    insert_headers = []
    insert_data = []
    entry_data = []

    if 'entry' in save_headers:
        entry_index = save_headers.index('entry')

        for header in save_headers:
            if header != 'entry':
                insert_headers.append(header)

        for row in save_data:
            entry_data.append(row[entry_index])
            row.pop(entry_index)
            insert_data.append(row)

        insert_column_string = ', '.join([header for header in insert_headers])
        update_column_string = ', '.join([header + ' = ?' for header in insert_headers])
        value_string = ', '.join(['?'] * len(insert_headers))

        for n in range(0, len(insert_data), 1):

            entry = entry_data[n]

            cursor.execute(f'SELECT * FROM {table_name}')
            if len(cursor.fetchall()) == 0 or entry in [None, '']:
                query = f'INSERT INTO {table_name} ({insert_column_string}) VALUES ({value_string})'
                print(query)
                cursor.execute(query, tuple(insert_data[n]))
            else:
                condition = f'entry = {entry}'
                query = f'UPDATE {table_name} SET {update_column_string} WHERE {condition}'
                print(query)
                cursor.execute(query, tuple(insert_data[n]))
    else:
        for header in save_headers:
            insert_headers.append(header)
        insert_column_string = ', '.join([header for header in insert_headers])
        value_string = ', '.join(['?'] * len(insert_headers))
        query = f'INSERT INTO {table_name} ({insert_column_string}) VALUES ({value_string})'
        for row in save_data:
            insert_data.append(row)

        cursor.executemany(query, tuple(insert_data))

    connection.commit()
    cursor.close()
    connection.close()

def is_number(number):
    try:
        float(number)
        return True
    except:
        return False

def is_date(date):
    try:
        datetime.datetime.strptime(date, '%Y-%m-%d')
        return True
    except:
        try:
            datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
            return True
        except:
            return False

def convert_to_capital_letter(number):
    letters = []

    while number > 0:
        # Get the remainder and convert it to a letter
        remainder = (number - 1) % 26
        letter = chr(remainder + ord('A'))

        # Add the letter to the list
        letters.append(letter)

        # Update the number for the next iteration
        number = (number - 1) // 26

    # Reverse the list and join the letters
    letters.reverse()
    result = ''.join(letters)

    return result


def find_closest_match(string, lists):
    if type(string) != str:
        return None
    best_match = None
    best_ratio = 0
    string_lists = []
    for item in lists:
        if item == None:
            string_lists.append('')
        else:
            string_lists.append(item)

    for string_list in string_lists:
        ratio = difflib.SequenceMatcher(None, string, string_list).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = string_list

    return best_match

def validator_list(header, table_name):

    validate_list_temp = get_list(table_name, header, [], '')
    validate_list = [item[0] for item in validate_list_temp]

    return validate_list

def get_project_list(user_name):
    table_name = 'project'
    return_header = 'project_name'
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers = [col[1] for col in cursor.fetchall()]
    conditions = []
    if user_name != '' and user_name != None:
        for header in database_headers:
            conditions.append([header, 'LIKE', user_name])
    cursor.close()
    connection.close()
    project_list = search_and_return(table_name, conditions, return_header)
    return project_list

def search_and_return(table_name, conditions, return_header):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    condition_array = []

    if len(conditions) > 0:
        for condition in conditions:
            if condition[1] == 'LIKE':
                condition_array.append('"' + condition[0] + '"' + condition[1] + '"' + "%" + str(condition[2]) + "%" + '"')
            else:
                condition_array.append('"' + condition[0] + '"' + condition[1] + '"' + str(condition[2]) + '"')
        query = f'SELECT {return_header} FROM {table_name} WHERE {"OR".join(condition_array)}'
    else:
        query = f'SELECT {return_header} FROM {table_name}'
    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()
    connection.close()
    data_list = []
    for item in data:
        data_list.append(item[0])

    return data_list

def upload_multi_excel(file_field):
    dropdown_sheet_name = "dropdown_list_setting"
    datatype_sheet_name = "datatype_setting"
    # filename = secure_filename(file.filename)
    workbook = load_workbook(file_field, data_only=True)
    sheet_names = workbook.sheetnames
    index_headers = ['entry']
    row_table_value = []
    dropdown_value = []
    datatype_value = []
    erorr_messages = []
    sheet_names_array = []
    data_matrix = []
    remove_column_table_array = []
    remove_column_name_array = []
    rename_column_table_array = []
    rename_column_old_name_array = []
    rename_column_new_name_array = []
    remove_row_condition_array = []
    remove_row_table_array = []

    file_entry = file_field.filename.split('_')[-1].replace('.xlsm', '').replace('.xlsx', '')
    print(f'field entry = {file_entry}')
    message, error, last_entry = check_download_log(file_entry, 'uploaded')
    if error:
        erorr_messages.append(message)
        return erorr_messages
    print('pass upload validation')

    ################################################### row table##############################
    worksheet = workbook['row_table']
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        formula_row = []
        for cell in row:
            formula_row.append(cell)
        row_table_value.append(formula_row)

    row_table_value_transpose = numpy.transpose(numpy.array(row_table_value))
    row_table = row_table_value_transpose[0]
    org_row_index = row_table_value_transpose[1]
    reference_row_index = row_table_value_transpose[2]
    for m in range(0, len(row_table), 1):
        if reference_row_index[m] == '#REF!':
            remove_condition = [['entry', '=', org_row_index[m]]]
            #remove_from_database(remove_condition, row_table[m])
            remove_row_condition_array.append(remove_condition)
            remove_row_table_array.append(row_table[m])

    ###################################column table####################################
    worksheet = workbook['column_table']
    column_table_value = []
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        print(f'column table row = {row}')
        formula_row = []
        for cell in row:
            if type(cell) == str:
                formula_row.append(cell.replace(' ', '_'))
                print(f'cell = {cell}')
            else:
                formula_row.append(cell)
                print(f'cell = {cell}')
        column_table_value.append(formula_row)
        print('column table')
        print(column_table_value)

    column_table_value_transpose = numpy.transpose(numpy.array(column_table_value))
    column_table = column_table_value_transpose[0]
    org_column_name = column_table_value_transpose[1]
    reference_column_name = column_table_value_transpose[2]

    column_dict = {}
    for m in range(0, len(column_table), 1):

        if reference_column_name[m] == '#REF!':
            # remove header
            remove_column_table_array.append(column_table[m])
            remove_column_name_array.append(org_column_name[m])
            #remove_header(column_table[m], org_column_name[m])
        elif reference_column_name[m] != org_column_name[m]:
            # rename header
            rename_column_table_array.append(column_table[m])
            rename_column_old_name_array.append(org_column_name[m])
            rename_column_new_name_array.append(reference_column_name[m])

    print('printing rename columns')
    print(rename_column_new_name_array)
    #rename_header(column_table[m], org_column_name[m], reference_column_name[m])
    if all(element == None for element in rename_column_new_name_array) and len(rename_column_new_name_array) > 0:
        erorr_messages.append('Excel file has not been saved before uploading')

    ################################################################################

    for sheet_name in sheet_names:
        print(f'sheet name = {sheet_name}')
        if sheet_name != dropdown_sheet_name and sheet_name != datatype_sheet_name and sheet_name != "row_table" and sheet_name != "column_table":
            worksheet = workbook[sheet_name]
            fetched_data = []
            for row in worksheet.iter_rows(min_row=1, values_only=True):
                fetched_data.append(row)

            # process the workbook here
            data_to_be_saved = fetched_data[1:]
            conditions = []
            saved_headers = list(fetched_data[0])

            for n in range(0, len(saved_headers), 1):
                if type(saved_headers[n]) == str:
                    saved_headers[n] = saved_headers[n].replace(' ', '_')

            for row in data_to_be_saved:
                condition = []
                for n in range(0, len(index_headers), 1):
                    condition.append([index_headers[n], '=', row[saved_headers.index(index_headers[n])]])
                conditions.append(condition)

            data_to_be_saved_list = [list(row) for row in data_to_be_saved]

            header_tuple = [tuple(saved_headers)]
            data_to_be_saved_tuple = [tuple(row) for row in data_to_be_saved_list]
            #save_uploaded(sheet_name, header_tuple + data_to_be_saved_tuple)
            sheet_names_array.append(sheet_name)
            data_matrix.append(header_tuple + data_to_be_saved_tuple)

        elif sheet_name == dropdown_sheet_name:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=1, values_only=True):
                formula_row = []
                for cell in row:
                    if type(cell) == str:
                        formula_row.append(cell.replace(' ', '_'))
                    else:
                        formula_row.append(cell)
                dropdown_value.append(formula_row)

        elif sheet_name == datatype_sheet_name:
            print(f'sheet name is {sheet_name}')
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=1, values_only=True):
                formula_row = []
                for cell in row:
                    if type(cell) == str:
                        formula_row.append(cell.replace(' ', '_'))
                    else:
                        formula_row.append(cell)
                datatype_value.append(formula_row)

    workbook = load_workbook(file_field)

    worksheet = workbook[dropdown_sheet_name]
    dropdown_formula = []
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        formula_row = []
        for cell in row:
            formula_row.append(cell)
        dropdown_formula.append(formula_row)
    dropdown_setting_table = []

    worksheet = workbook[datatype_sheet_name]
    datatype_formula = []
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        formula_row = []
        for cell in row:
            formula_row.append(cell)
        datatype_formula.append(formula_row)
    datatype_setting_table = []

    for row in range(1, len(dropdown_value), 1):
        data_row = []
        for col in range(0, len(dropdown_value[row]), 1):
            print(f'dropdown_formula = {dropdown_formula[row][col]}')
            if type(dropdown_formula[row][col]) == str and '!' in dropdown_formula[row][col]:
                index = dropdown_formula[row][col].index('!')
                ref_table_name = dropdown_formula[row][col][:index].replace('=', '')
                ref_column_name = str(dropdown_value[row][col])
                combined_ref_string = ref_table_name + '>>' + ref_column_name
                data_row.append(combined_ref_string)
            else:
                data_row.append(dropdown_formula[row][col])
        dropdown_setting_table.append(data_row)
    header_tuple = [tuple(dropdown_formula[0])]
    data_to_be_saved_tuple = [tuple(row) for row in dropdown_setting_table]
    sheet_names_array.append(dropdown_sheet_name)
    data_matrix.append(header_tuple + data_to_be_saved_tuple)

    print(f'datatype value = {datatype_value}')
    for row in range(1, len(datatype_value), 1):
        print('datatype processing')
        data_row = []
        for col in range(0, len(datatype_value[row]), 1):
            print(f'datatype_formula = {datatype_formula[row][col]}')
            if type(datatype_formula[row][col]) == str and '!' in datatype_formula[row][col]:
                index = datatype_formula[row][col].index('!')
                ref_table_name = datatype_formula[row][col][:index].replace('=', '')
                ref_column_name = str(datatype_value[row][col])
                combined_ref_string = ref_table_name + '>>' + ref_column_name
                data_row.append(combined_ref_string)
            else:
                data_row.append(datatype_formula[row][col])
        datatype_setting_table.append(data_row)
    header_tuple = [tuple(datatype_formula[0])]
    data_to_be_saved_tuple = [tuple(row) for row in datatype_setting_table]
    sheet_names_array.append(datatype_sheet_name)
    data_matrix.append(header_tuple + data_to_be_saved_tuple)

    for k in range(0, len(sheet_names_array), 1):
        '''if sheet_names_array[k] == 'time_sheet':
            if 'user_name' not in data_matrix[k][0]:
                erorr_messages.append('The header "user name" is not found in the Tab "time_sheet". It has either been deleted or renamed. Please make sure this header is included in the database.')
            if 'submit_date' not in data_matrix[k][0]:
                erorr_messages.append('The header "submit date" is not found in the Tab "time_sheet". It has either been deleted or renamed. Please make sure this header is included in the database.')
            if 'status' not in data_matrix[k][0]:
                erorr_messages.append(
                    'The header "status" is not found in the Tab "time_sheet". It has either been deleted or renamed. Please make sure this header is included in the database.')
            if 'hours' not in data_matrix[k][0]:
                erorr_messages.append(
                    'The header "hours" is not found in the Tab "time_sheet". It has either been deleted or renamed. Please make sure this header is included in the database.')

        if sheet_names_array[k] == 'user':
            if 'user_name' not in data_matrix[k][0]:
                print('user_name not in user')
                erorr_messages.append('The header "user name" is not found in the Tab "user". It has either been deleted or renamed. Please make sure this header is included in the database.')
            if 'first_name' not in data_matrix[k][0]:
                erorr_messages.append('The header "first name" is not found in the Tab "user". It has either been deleted or renamed. Please make sure this header is included in the database.')
            if 'last_name' not in data_matrix[k][0]:
                erorr_messages.append('The header "last name" is not found in the Tab "user". It has either been deleted or renamed. Please make sure this header is included in the database.')'''

        if sheet_names_array[k] == 'datatype_setting':
            special_type_index = data_matrix[k][0].index('special_type')
            special_types = ['timesheet user', 'username', 'email', 'submit date', 'work hours', 'project id', 'client id', 'discipline id']
            for special_type in special_types:
                count = 0
                for n in range(1, len(data_matrix[k]), 1):
                    print(f'{special_type} ?= {data_matrix[k][n][special_type_index]}')

                    if data_matrix[k][n][special_type_index] == special_type:
                        count = count + 1
                        print('matched')
                if count == 0:
                    erorr_messages.append(f'{special_type} has not been assigned in datatype setting tab')
                elif count > 1:
                    erorr_messages.append(f'{special_type} cannot be assigned more than once in datatype setting tab')

    if len(erorr_messages) > 0:
        print('error message length > 0')
        return erorr_messages

    else:

        for k in range(0, len(remove_row_table_array), 1):
            remove_from_database(remove_row_condition_array[k], remove_row_table_array[k])

        for k in range(0, len(remove_column_table_array), 1):
            remove_header(remove_column_table_array[k], remove_column_name_array[k])

        for k in range(0, len(rename_column_table_array), 1):
            rename_header(rename_column_table_array[k], rename_column_old_name_array[k], rename_column_new_name_array[k])

        for k in range(0, len(sheet_names_array), 1):
            save_uploaded(sheet_names_array[k], data_matrix[k])

        return erorr_messages

def upload_excel(table_name, index_headers, file_field, check_headers=[], check_tables=[]):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA table_info({table_name})')

    header_info = cursor.fetchall()
    cursor.close()
    connection.close()

    # filename = secure_filename(file.filename)
    workbook = load_workbook(file_field)
    worksheet = workbook.active
    fetched_data = []
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        fetched_data.append(row)

    # process the workbook here

    data_to_be_saved = fetched_data[1:]

    conditions = []
    saved_headers = fetched_data[0]
    for n in range(0, len(saved_headers), 1):
        saved_headers[n] = saved_headers[n].replace(' ', '_')

    for row in data_to_be_saved:
        condition = []
        for n in range(0, len(index_headers), 1):
            condition.append([index_headers[n], '=', row[saved_headers.index(index_headers[n])]])
        conditions.append(condition)


    data_to_be_saved_list = [list(row) for row in data_to_be_saved]
    error_messages = []

    for n in range(0, len(check_headers), 1):
        if type(check_tables[n]) == str:
            validate_list = validator_list(check_headers[n], check_tables[n])
        elif type(check_tables[n]) == list:
            validate_list = check_tables[n]
        else:
            validate_list = None
        header_index = saved_headers.index(check_headers[n])
        for m in range(0, len(data_to_be_saved), 1):
            column_index = saved_headers.index(check_headers[n])
            info_index = [row[1] for row in header_info].index(check_headers[n])

            string_to_match = data_to_be_saved_list[m][header_index]
            if type(validate_list) == list:
                data_to_be_saved_list[m][header_index] = find_closest_match(string_to_match, validate_list)

            if data_to_be_saved_list[m][header_index] == None:
                error_messages.append(f'Error uploading. Cannot find match for input "{string_to_match}" on Cell {convert_to_capital_letter(column_index + 1)}{m + 2}')

            if header_info[info_index][2] == 'DATE' and not is_date(string_to_match):
                error_messages.append(f'Error uploading. Date format in "{string_to_match}" on Cell {convert_to_capital_letter(column_index + 1)}{m + 2} must be "yyyy-mm-dd" or "yyyy-mm-dd hh:mm:ss"')

            if header_info[info_index][2] in ['REAL', 'INTEGER'] and not is_number(string_to_match):
                error_messages.append(f'Error uploading. "{string_to_match}" on Cell {convert_to_capital_letter(column_index + 1)}{m + 2} should be integer or float')

            if header_info[info_index][2] in ['REAL'] and is_number(string_to_match):
                data_to_be_saved_list[m][header_index] = float(string_to_match)

            if header_info[info_index][2] in ['INTEGER'] and is_number(string_to_match):
                data_to_be_saved_list[m][header_index] = int(string_to_match)



    if len(error_messages) == 0:
        header_tuple = [tuple(saved_headers)]
        data_to_be_saved_tuple = [tuple(row) for row in data_to_be_saved_list]
        save_uploaded(table_name, header_tuple + data_to_be_saved_tuple)
        #save_table(table_name, conditions, data_to_be_saved_tuple, saved_headers)

    return error_messages

def delete_row(table_name, conditions):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    condition_string = query_conditions(conditions)

    cursor.execute(f'DELETE FROM "{table_name}" WHERE {condition_string}')
    connection.commit()
    cursor.execute(f'SELECT * FROM {table_name}')

    connection.close()

def save_table(table_name, conditions, data_to_be_saved, saved_headers):
    print('saved headers')
    print(saved_headers)
    print('data to be saved')
    print(data_to_be_saved)

    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers = [col[1] for col in cursor.fetchall()]
    ################## insert columnes if missing ############
    for header in saved_headers:
        if header not in database_headers:
            cursor.execute(f'ALTER TABLE {table_name} ADD COLUMN "{header}"')
    #############################################################

    for n in range(0, len(data_to_be_saved), 1):

        ############## get row ###############################
        row = []
        mod_headers = []
        for m in range(0, len(data_to_be_saved[n]), 1):
            if data_to_be_saved[n][m] != None and saved_headers[m] != 'entry':
                row.append(data_to_be_saved[n][m])
                mod_headers.append(saved_headers[m])

        ############### get query ######################
        update_column_string = ', '.join(['"' + header + '" = ?' for header in mod_headers])
        insert_column_string = ', '.join(['"' + header + '"' for header in mod_headers])
        value_string = ', '.join(['?'] * len(mod_headers))

        condition_string = query_conditions(conditions[n])
        ##########################################################
        if condition_string != '':
            ############## check to see if row exist #################

            cursor.execute(f'SELECT * FROM "{table_name}" WHERE {condition_string}')
            result = cursor.fetchall()
        else:
            result = []

        if len(result) > 0:
            row_exist = True
        else:
            row_exist = False

        ########################################################


        if row_exist:
            query = f'UPDATE "{table_name}" SET {update_column_string} WHERE {condition_string}'
            cursor.execute(query, tuple(row))
        else:
            query = f'INSERT INTO "{table_name}" ({insert_column_string}) VALUES ({value_string})'

            cursor.execute(query, tuple(row))

    connection.commit()
    cursor.close()
    connection.close()

def get_list(table_name, id_header, display_headers, concat_string):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()
    headers = [id_header]
    headers.extend(display_headers)

    headers_quoted = ['"' + header + '"' for header in headers]
    headers_string = ', '.join(headers_quoted)
    query = f'SELECT {headers_string} FROM {table_name}'

    cursor.execute(query)

    fetched_data = cursor.fetchall()
    choice_list = []

    for item in fetched_data:
        display_holder = []
        for n in range(1, len(item), 1):
            if item[n] == None:
                display_holder.append((''))
            else:
                display_holder.append(item[n])
        choice_list.append((item[0], concat_string.join(display_holder)))
    cursor.close()
    connection.close()
    return choice_list

def query_conditions(conditions, and_or=None):
    conditions_array = []

    for condition in conditions:

        '''if condition[0] == 'entry' and condition[1] == '=' and condition[2] in [None, '']:
            pass'''
        if type(condition[2]) == datetime.datetime:
            cond3 = f'strftime("%Y-%m-%d", "{condition[2]}")'
            cond1 = f'strftime("%Y-%m-%d", "{condition[0]}")'
            cond2 = condition[1]
            conditions_array.append(cond1 + ' ' + cond2 + ' ' + cond3)
        elif type(condition[2]) == int:
            cond3 = f'CAST("{condition[2]}" AS INTEGER)'
            cond1 = f'CAST("{condition[0]}" AS INTEGER)'
            cond2 = condition[1]
            conditions_array.append(cond1 + ' ' + cond2 + ' ' + cond3)
        elif type(condition[2]) == list:
            in_arr = ['"' + item + '"' for item in condition[2]]
            cond3 = '(' + ', '.join(in_arr) + ')'
            cond1 = '"' + condition[0] + '"'
            cond2 = condition[1]
            conditions_array.append(cond1 + ' ' + cond2 + ' ' + cond3)
        elif condition[2] == None:
            cond3 = '"' + '"'
            cond1 = '"' + condition[0] + '"'
            cond2 = condition[1]
            conditions_array.append(cond1 + ' ' + cond2 + ' ' + cond3)

        else:
            cond3 = '"' + condition[2] + '"'
            cond1 = '"' + condition[0] + '"'
            cond2 = condition[1]
            conditions_array.append(cond1 + ' ' + cond2 + ' ' + cond3)
    ############################################
    if and_or == None or and_or == 'and':
        query_string = ' AND '.join(conditions_array)
    elif and_or == 'or':
        query_string = ' OR '.join(conditions_array)
    ###########################################

    return query_string

def download_multi_data(table_names, checked_headers, conditions):

    column_table_array = []
    column_org_header_array = []
    reference_column_array = []

    row_table_array = []
    row_index_array = []
    reference_row_array = []

    _, _, last_entry = check_download_log(None, 'downloaded')

    dropdown_sheet_name = 'dropdown_list_setting'
    output = io.BytesIO()
    #writer = pandas.ExcelWriter(output, engine='openpyxl')
    wb = load_workbook('excel_template.xlsm', keep_vba=True)

    for n in range(0, len(table_names), 1):

        headers = checked_headers[n]

        if len(headers) > 0:

            headers.insert(0, 'entry')
            headers_string = ','.join(headers)
            query = f'SELECT {headers_string} FROM {table_names[n]}'
            query_string = query_conditions(conditions[n])
            if len(query_string) > 0:
                query = query + ' where ' + query_string

            print(f'table name is {table_names[n]}')
            table_content = download_table(table_names[n], headers)
            if table_names[n] not in ['dropdown_list_setting', 'datatype_setting']:
                populate_excel(wb, table_names[n], table_content)

            #column_index_array.extend(get_header_index(table_names[n], headers))
            for m in range(0, len(table_content[0]), 1):
                column_table_array.append(table_names[n])
                column_org_header_array.append(table_content[0][m])
                #column_index_array.append(table_content[0][m])
                reference_column_array.append('=' + table_names[n] + '!' + column_number_to_letter(m + 1) + '1')

            entry_column_number = table_content[0].index('entry') + 1
            for m in range(1, len(table_content), 1):
                row_table_array.append(table_names[n])
                row_index_array.append(table_content[m][entry_column_number - 1])
                reference_row_array.append('=' + table_names[n] + '!' + column_number_to_letter(entry_column_number) + str(m + 1))

    column_table = numpy.transpose(numpy.array([column_table_array, column_org_header_array, reference_column_array]))

    row_table = numpy.transpose(numpy.array([row_table_array, row_index_array, reference_row_array]))
    populate_excel(wb, 'column_table', column_table)
    populate_excel(wb, 'row_table', row_table)

    ################download dropdown setting###########################
    #query = f'SELECT target, source, criteria_1, filter_1, criteria_2, filter_2, criteria_3, filter_3, criteria_4, filter_4 FROM dropdown_list_setting'
    dropdown_headers = ['entry', 'target', 'source', 'criteria_1', 'filter_1', 'criteria_2', 'filter_2', 'criteria_3', 'filter_3', 'criteria_4', 'filter_4', 'dropdown_list_type']
    dropdown_data_org = download_table(dropdown_sheet_name, dropdown_headers)

    dropdown_formula = [dropdown_headers]

    for row in range(1, len(dropdown_data_org), 1):
        dropdown_formula.append([])
        for col in range(0, len(dropdown_data_org[row]), 1):

            if type(dropdown_data_org[row][col]) == str and dropdown_data_org[row][col] is not None and '>>' in dropdown_data_org[row][col]:
                sheet_name = dropdown_data_org[row][col].split('>>')[0]
                column_name = dropdown_data_org[row][col].split('>>')[1]

                if sheet_name in table_names:
                    table_index = table_names.index(sheet_name)
                    if column_name in checked_headers[table_index]:
                        column_digit = checked_headers[table_index].index(column_name) + 1

                        column_letter = column_number_to_letter(column_digit)
                        formula = '=' + sheet_name + '!' + column_letter + '$1'
                        dropdown_formula[-1].append(formula)
                    else:
                        dropdown_formula[-1].append(f'missing Column {column_name} from Table {sheet_name}')
                else:
                    dropdown_formula[-1].append(f'missing Table {sheet_name}')
            else:
                dropdown_formula[-1].append(dropdown_data_org[row][col])
    populate_excel(wb, dropdown_sheet_name, dropdown_formula)

    #######################datatype_setting########################################
    datatype_sheet_name = 'datatype_setting'
    datatype_headers = ['entry', 'column_reference', 'datatype', 'uniqueness', 'special_type']
    datatype_data_org = download_table(datatype_sheet_name, datatype_headers)
    datatype_formula = [datatype_headers]

    for row in range(1, len(datatype_data_org), 1):
        datatype_formula.append([])
        for col in range(0, len(datatype_data_org[row]), 1):

            if type(datatype_data_org[row][col]) == str and datatype_data_org[row][col] is not None and '>>' in datatype_data_org[row][col]:
                sheet_name = datatype_data_org[row][col].split('>>')[0]
                column_name = datatype_data_org[row][col].split('>>')[1]

                if sheet_name in table_names:
                    table_index = table_names.index(sheet_name)
                    if column_name in checked_headers[table_index]:
                        column_digit = checked_headers[table_index].index(column_name) + 1

                        column_letter = column_number_to_letter(column_digit)
                        formula = '=' + sheet_name + '!' + column_letter + '$1'
                        datatype_formula[-1].append(formula)
                    else:
                        datatype_formula[-1].append(f'missing Column {column_name} from Table {sheet_name}')
                else:
                    datatype_formula[-1].append(f'missing Table {sheet_name}')
            else:
                datatype_formula[-1].append(datatype_data_org[row][col])
    populate_excel(wb, datatype_sheet_name, datatype_formula)

    #df = pandas.DataFrame(dropdown_formula)
    #df.to_excel(writer, sheet_name=dropdown_sheet_name, index=False, header=False)
    ######################################################################
    #writer.close()

    wb.remove(wb['Introduction'])

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm')
    temp_filename = temp_file.name
    wb.save(temp_filename)
    temp_file.close()

    # Create a response with the Excel file
    response = make_response(open(temp_filename, 'rb').read())
    response.headers.set('Content-Disposition', 'attachment', filename=f'database_{last_entry + 1}.xlsm')
    response.headers.set('Content-Type', 'application/vnd.ms-excel.sheet.macroEnabled.12')

    # Delete the temporary file
    os.remove(temp_filename)

    return response


def download_data(table_name, conditions):
    connection = sqlite3.connect('data_base.db')
    cursor = connection.cursor()

    query = f'SELECT * FROM {table_name}'
    query_string = query_conditions(conditions)

    if len(query_string) > 0:
        query = query + ' where ' + query_string

    cursor.execute(query)

    ############ save file #######
    output = io.BytesIO()
    df = pandas.read_sql_query(query, connection)
    writer = pandas.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.close()
    ############

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=time_sheet.xlsx'

    cursor.close()
    connection.close()

    return response

def get_database_table(table_name, conditions, selected_headers, optional_header):
    user_data = []
    optional_data = []
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA table_info({table_name})')
    headers = [col[1] for col in cursor.fetchall()]

    selected_headers_index = [headers.index(header) for header in selected_headers]
    optional_index = headers.index(optional_header)

    query = f'SELECT * FROM {table_name}'

    query_string = query_conditions(conditions)

    if len(query_string) > 0:
        query = query + ' where ' + query_string

    cursor.execute(query)
    fetched_data = cursor.fetchall()

    for array in fetched_data:
        user_data.append([])
        optional_data.append(array[optional_index])
        for index in selected_headers_index:
            user_data[-1].append(array[index])
    cursor.close()
    connection.close()

    return user_data, optional_data

def remove_from_database(conditions, table_name):
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    #remove_keys = remove_condition.keys()
    remove_condition_array = []

    if len(conditions) > 0:
        for condition in conditions:
            remove_condition_array.append('"' + condition[0] + '"' + condition[1] + '"' + str(condition[2]) + '"')
        query = f'DELETE FROM {table_name} WHERE {"AND".join(remove_condition_array)}'
    else:
        query = f'SELECT * FROM {table_name}'


    cursor.execute(query)
    connection.commit()
    cursor.close()
    connection.close()

def check_existing(table_name, conditions):
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    table_exist = cursor.fetchone()
    #keys = condition.keys()
    condition_array = []

    for condition in conditions:

        condition_array.append('"' + condition[0] + '"' + condition[1] + '"' + condition[2] + '"')

    query = f'SELECT EXISTS(SELECT 1 FROM {table_name} WHERE {"AND".join(condition_array)})'


    row_exist = False
    if table_exist:
        cursor.execute(query)
        row_exist = cursor.fetchone()[0]
    cursor.close()
    connection.close()

    return row_exist

def check_username(username):
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    table_name = 'users'

    cursor.execute(f'SELECT EXISTS(SELECT 1 from {table_name} where username = ?)', (username,))
    existence = cursor.fetchone()[0]
    cursor.close()
    connection.close()
    return existence

def fill_form_list(form_list, match_header, table_name):

    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers = [col[1] for col in cursor.fetchall()]

    fetched_rows = []

    for row in form_list.table_form.entries:

        for field in row:

            splitname = field.name.split('-')
            if len(splitname) >= 3:
                fieldname = '-'.join(splitname[2:])
            else:
                fieldname = field.name

            if fieldname == match_header:

                query = f'SELECT * FROM {table_name} WHERE {fieldname} = "{field.data}"'
                print(f'query fill list = {query}')
                cursor.execute(query)

                fetched_rows.append(cursor.fetchone())
                print(fetched_rows[-1])

    n = 0
    for row in form_list.table_form.entries:
        for field in row:
            splitname = field.name.split('-')
            if len(splitname) >= 3:
                fieldname = '-'.join(splitname[2:])
            else:
                fieldname = field.name
            if fieldname in database_headers and fieldname != 'entry':
                index = database_headers.index(fieldname)
                if field.data == None or field.data == '':
                    field.data = fetched_rows[n][index]
        n += 1

def load_form(form, conditions, table_name):
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()

    #load_keys = conditions.keys()
    load_condition_array = []

    if len(conditions) > 0:
        for condition in conditions:
            load_condition_array.append('"' + condition[0].replace(' ', '_') + '"' + condition[1] + '"' + condition[2] + '"')

        query = f'SELECT * FROM "{table_name}" WHERE {"AND".join(load_condition_array)}'
    else:
        query = f'SELECT * FROM {table_name}'

    cursor.execute(query)
    fetched_data = cursor.fetchall()

    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers = [col[1] for col in cursor.fetchall()]

    for field in form:
        if field.label.text.lower().replace(' ', '_') in database_headers:
            index = database_headers.index(field.label.text.lower().replace(' ', '_'))
            value = fetched_data[0][index]
            if type(value) == str:
                value = value.split(',')
                if len(value) == 1:
                    value = value[0]
            try:
                value = datetime.datetime.strptime(value, '%Y-%m-%d')

            except:
                pass

            field.data = value

    cursor.close()
    connection.close()

def load_dynamic_form(form, conditions, table_name, form_status):
    print(f'dynamic form condition = {conditions}')
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    form_header = []

    ################# check to see if form headers matches with database ################
    fieldnames = []
    #form_list.table_form.append_entry()
    for field in form:
        splitname = field.name.split('-')
        if len(splitname) >= 3:
            fieldname = '-'.join(splitname[2:])
        else:
            fieldname = field.name
        fieldnames.append(fieldname)
    print(f'initial fieldnames = {fieldnames}')
    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    database_headers_info = cursor.fetchall()

    database_headers = [row[1] for row in database_headers_info]
    database_headers_datatype = [row[2] for row in database_headers_info]
    class_obj = type(form)


    '''######### determin the index number of the entry #############
    entry_index = None
    for condition in conditions:
        if condition[0] == 'entry':
            entry_index = int(condition[1])
    ##############################################################'''

    '''for n in range(0, len(database_headers), 1):

        if database_headers[n] not in fieldnames:
            #for row in form_list.table_form.entries:
            print('calling select field type')
            select_type = select_field_type(table_name, database_headers[n])
            if select_type == 'single_select':
                setattr(class_obj, database_headers[n], SelectField(label=database_headers[n]))
            elif select_type == 'multi_select':
                setattr(class_obj, database_headers[n], SelectMultipleField(label=database_headers[n]))
            elif database_headers_datatype[n] == 'INTEGER':
                setattr(class_obj, database_headers[n], IntegerField(label=database_headers[n]))
            elif database_headers_datatype[n] == 'REAL':
                setattr(class_obj, database_headers[n], DecimalField(label=database_headers[n]))
            elif find_column_datatype(table_name, database_headers[n]) == 'date':
                print('date field found in form')
                setattr(class_obj, database_headers[n], DateField(label=database_headers[n]))
            else:
                setattr(class_obj, database_headers[n], StringField(label=database_headers[n]))'''

    for field in form:
        form_header.append(field.label.text)

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    table_exist = cursor.fetchone()

    if table_exist:
        if len(conditions) > 0:
            query_condition = query_conditions(conditions)
            query = f'SELECT * FROM {table_name} WHERE {query_condition}'

        else:
            query = f'SELECT * FROM {table_name}'
    else:
        return

    cursor.execute(query)
    fetched_data = cursor.fetchone()

    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers = [col[1] for col in cursor.fetchall()]
    entry_column = database_headers.index('entry')

    '''for _ in fetched_data:
        form_list.table_form.append_entry()'''

    #n = 0
    #for form in form_list.table_form.entries:
    for field in form:

        splitname = field.name.split('-')
        if len(splitname) >= 3:
            fieldname = '-'.join(splitname[2:])
        else:
            fieldname = field.name

        if fieldname in database_headers:

            index = database_headers.index(fieldname)

            #value = fetched_data[n][index]
            if fetched_data != None:

                if field.data not in [None] and field.data != fetched_data[index]:
                    value = field.data
                else:
                    value = fetched_data[index]
            else:
                value = field.data

            if type(value) == str:
                value = value.split(',')
                if len(value) == 1:
                    value = value[0]
            try:
                value = datetime.datetime.strptime(value, '%Y-%m-%d')

            except:
                pass

            print(f'field = {field.name} data type = {field.type}')
            if field.type == 'SelectField':
                if fetched_data != None:
                    dropdown_list = get_dropdown_list(table_name, fieldname, fetched_data[entry_column])
                    print(f'table name = {table_name}')
                    print(f'form field = {field.name}')
                    print(f'dropdown list = {dropdown_list}')
                    print(f'dropdown value = {value}')

                    field.choices = dropdown_list
                    field.data = value
                else:
                    database_values = []
                    for _ in database_headers:
                        database_values.append([])
                    dropdown_list = get_dropdown_list(table_name, fieldname, None, headers=database_headers, cell_values=database_values)
                    field.choices = dropdown_list
                    field.data = value
            elif field.type == 'SelectMultipleField':
                if fetched_data != None:
                    dropdown_list = get_dropdown_list(table_name, fieldname, fetched_data[entry_column])
                    field.choices = dropdown_list
                    print(f'multiselect fieldname = {field.name}')
                    print(f'multiselect field value = {value}')
                    field.data = value
                    #if type(value) == str:
                        #value_list_comma = value.split(',')
                        #field.data = value_list_comma

                    print(f'dropdown list = {dropdown_list}')
                    print(f'multifield value = {value}')
                else:
                    database_values = []
                    for _ in database_headers:
                        database_values.append([])
                    dropdown_list = get_dropdown_list(table_name, fieldname, None, headers=database_headers, cell_values=database_values)
                    field.choices = dropdown_list
                    field.data = value
            else:
                field.data = value

        #n += 1

    cursor.close()
    connection.close()

    ###################### testing #####################
    '''for formfield in form_list.table_form[-1]:
        print(f'field names == {formfield.name}, value = {formfield.data}')'''
    ###################################################################################
    if form_status == 'new' and 'password' in database_headers:
        password_index = database_headers.index('password')
        database_headers.insert(password_index + 1, 'confirm_password')
    '''elif form_status == 'edit' and 'password' in database_headers:
        database_headers.remove('password')'''

    return database_headers

def load_dynamic_form_list(form_list, conditions, table_name, and_or=None):
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    form_header = []
    ################# check to see if form headers matches with database ################
    fieldnames = []
    form_list.table_form.append_entry()

    for field in form_list.table_form[0]:
        splitname = field.name.split('-')
        if len(splitname) >= 3:
            fieldname = '-'.join(splitname[2:])
        else:
            fieldname = field.name
        fieldnames.append(fieldname)

    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    database_headers_info = cursor.fetchall()

    database_headers = [row[1] for row in database_headers_info]
    database_headers_datatype = [row[2] for row in database_headers_info]
    class_obj = form_list.table_form.unbound_field.args[0]

    for n in range(0, len(database_headers), 1):
        #if database_headers[n] not in fieldnames:
        for _ in form_list.table_form.entries:
            select_type = select_field_type(table_name, database_headers[n])
            print(f'table name = {table_name}, header = {database_headers[n]}, field type = {select_type}')
            if select_type == 'single_select':
                setattr(class_obj, database_headers[n], SelectField(label=database_headers[n]))
            elif select_type == 'multi_select':
                setattr(class_obj, database_headers[n], SelectMultipleField(label=database_headers[n]))
            elif find_column_datatype(table_name, database_headers[n]) == 'integer':
                setattr(class_obj, database_headers[n], IntegerField(label=database_headers[n]))
            elif find_column_datatype(table_name, database_headers[n]) == 'number':
                setattr(class_obj, database_headers[n], DecimalField(label=database_headers[n]))
            elif find_column_datatype(table_name, database_headers[n]) == 'date' and database_headers[n] != 'submit_date':
                print('date field found in form')
                setattr(class_obj, database_headers[n], DateField(label=database_headers[n]))
            else:
                setattr(class_obj, database_headers[n], StringField(label=database_headers[n]))

    try:
        while True:
            form_list.table_form.pop_entry()
    except:
        pass

    form_list.table_form.append_entry()

    ###################### testing #####################
    '''for formfield in form_list.table_form[-1]:
        print(f'field names == {formfield.name}, value = {formfield.data}')'''
    ###################################################################################

    for form in form_list.table_form.entries:
        for field in form:
            form_header.append(field.label.text)
        break
    form_list.table_form.pop_entry()

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    table_exist = cursor.fetchone()

    if table_exist:
        if len(conditions) > 0:

            query_condition = query_conditions(conditions, and_or=and_or)

            query = f'SELECT * FROM {table_name} WHERE {query_condition}'
        else:
            query = f'SELECT * FROM {table_name}'
    else:
        return

    print(f'query = {query}')
    cursor.execute(query)
    fetched_data = cursor.fetchall()


    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers = [col[1] for col in cursor.fetchall()]
    print(database_headers)
    entry_column = database_headers.index('entry')
    for _ in fetched_data:
        form_list.table_form.append_entry()
    n = 0
    for form in form_list.table_form.entries:
        for field in form:
            splitname = field.name.split('-')
            if len(splitname) >= 3:
                fieldname = '-'.join(splitname[2:])
            else:
                fieldname = field.name
            if fieldname in database_headers:
                index = database_headers.index(fieldname)
                #print(f'fieldnmae = {field.name}, table value = {fetched_data[n][index]}')
                #value = fetched_data[n][index]
                if fetched_data != None:
                    value = fetched_data[n][index]
                else:
                    value = field.data

                try:
                    value = datetime.datetime.strptime(value, '%Y-%m-%d')
                except:
                    pass

                if field.type == 'SelectField':
                    print(f'table_name = {table_name}')
                    print(f'fieldname = {fieldname}')
                    print(f'fetched data = {fetched_data[n][entry_column]}')
                    dropdown_list = get_dropdown_list(table_name, fieldname, fetched_data[n][entry_column])
                    print(f'dropdown list = {dropdown_list}')
                    field.choices = dropdown_list
                field.data = value

        n += 1

    cursor.close()
    connection.close()

    return database_headers

def load_form_list(form_list, conditions, table_name):
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    form_header = []

    try:
        while True:
            form_list.table_form.pop_entry()
    except:
        pass

    form_list.table_form.append_entry()
    for form in form_list.table_form.entries:
        for field in form:
            form_header.append(field.label.text)
        break
    form_list.table_form.pop_entry()

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    table_exist = cursor.fetchone()

    if table_exist:
        if len(conditions) > 0:
            query_condition = query_conditions(conditions)
            query = f'SELECT * FROM {table_name} WHERE {query_condition}'
        else:
            query = f'SELECT * FROM {table_name}'
    else:
        return

    cursor.execute(query)
    fetched_data = cursor.fetchall()

    cursor.execute(f'PRAGMA table_info({table_name})')
    database_headers = [col[1] for col in cursor.fetchall()]

    for _ in fetched_data:
        form_list.table_form.append_entry()

    n = 0
    for form in form_list.table_form.entries:
        for field in form:

            splitname = field.name.split('-')
            if len(splitname) >= 3:
                fieldname = '-'.join(splitname[2:])
            else:
                fieldname = field.name

            if fieldname in database_headers:

                index = database_headers.index(fieldname)

                value = fetched_data[n][index]
                if type(value) == str:
                    value = value.split(',')
                    if len(value) == 1:
                        value = value[0]
                try:
                    value = datetime.datetime.strptime(value, '%Y-%m-%d')

                except:
                    pass

                field.data = value

        n += 1

    cursor.execute(f'PRAGMA TABLE_INFO({table_name})')
    header_info = cursor.fetchall()

    cursor.close()
    connection.close()

    return header_info

def save_form(form_list, target_condition, table_name):
    connection = sqlite3.connect(db_file)
    cursor = connection.cursor()
    full_headers = []
    full_headers_type = []
    full_headers_data = []
    for form in form_list:
        print(f'form data = {form.data}')
        if type(form) not in [SubmitField] and form.label.text.lower().replace(' ', '_') != 'confirm_password' and form.label.text.lower().replace(' ', '_') != 'csrf_token':
            splitname = form.name.split('-')
            if len(splitname) >= 3:
                fieldname = splitname[2]
            else:
                fieldname = form.name
            print(f'save form header name = {fieldname}')
            '''label = form.label.text
            label = label.replace(' ', '_')
            label = label.lower()'''
            full_headers.append(fieldname)
            full_headers_type.append('text')
            print(f'save form header type = {full_headers_type[-1]}')

            if type(form.data) == list:
                full_headers_data.append(','.join(form.data))
            else:
                full_headers_data.append(form.data)

            if form.type == 'SelectField':
                print(f'select value = {form.choices}')
                pass
            print(f'save form header data = {full_headers_data[-1]}')
    full_headers_string_array = []
    for n in range(0, len(full_headers), 1):
        full_headers_string_array.append("'" + full_headers[n] + "'" + ' ' + full_headers_type[n])
    full_headers_string = '(' + ', '.join(full_headers_string_array) + ')'

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))

    if not cursor.fetchone():
        cursor.execute(f"create table '{table_name}' {full_headers_string}")
    else:
        cursor.execute(f'PRAGMA table_info({table_name})')
        existing_headers = [col[1] for col in cursor.fetchall()]

        for n in range(0, len(full_headers), 1):
            if full_headers[n].replace("'", '').replace('"', "") not in existing_headers:
                query = f"ALTER TABLE {table_name} ADD COLUMN {full_headers[n]} {full_headers_type[n]}"
                cursor.execute(query)

    cursor.execute(f'PRAGMA table_info({table_name})')
    updated_headers = [col[1] for col in cursor.fetchall()]
    connection.commit()
    cursor.close()
    connection.close()

    rearranged_data = []

    for header in updated_headers:
        if header.replace("'", "") in full_headers:
            header_index = full_headers.index(header.replace("'", ""))
            rearranged_data.append(full_headers_data[header_index])
        else:
            rearranged_data.append('')

    data_question_mark = ','.join(['?'] * len(rearranged_data))
    query = f"INSERT INTO {table_name} VALUES ({data_question_mark})"

    #cursor.executemany(f"INSERT INTO {table_name} VALUES ({data_question_mark})", [tuple(rearranged_data)])
    condition = []
    condition.append(target_condition)

    save_table(table_name, condition, [tuple(rearranged_data)], updated_headers)
